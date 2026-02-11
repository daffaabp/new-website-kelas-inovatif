from openpyxl import load_workbook
import sys

try:
    wb = load_workbook(filename='semrush-issues/kelasinovatif.com_duplicate_title_tag_20260208.xlsx')
    sheet = wb.active
    
    print("Rows:")
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i < 20:  # Print first 20 rows
            print(row)
        else:
            break
except Exception as e:
    print(f"Error: {e}")
