from openpyxl import load_workbook
import sys

try:
    wb = load_workbook(filename='semrush-issues/kelasinovatif.com_duplicate_meta_description_20260208.xlsx')
    sheet = wb.active
    
    print("Rows (First 20):")
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i < 20:
            print(row)
        else:
            break
except Exception as e:
    print(f"Error: {e}")
